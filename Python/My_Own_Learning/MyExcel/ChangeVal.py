import openpyxl as xl

wb = xl.load_workbook('PriceList.xlsx')
sheet = wb['Sheet1']
for i in range(2, sheet.max_row + 1):
    ccell = sheet.cell(i, 4)
    cpr = ccell.value * 0.5
    ncell = sheet.cell(i, 5)
    ncell.value = cpr
    
wb.save('ZNewPriceList.xlsx')
    